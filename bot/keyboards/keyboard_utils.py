"""Утилиты для создания inline клавиатур."""
import json

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from bot.config import BASEDIR
from bot.constants import VIOLATION_CATEGORY_JSON_FILE


def create_inline_buttons_from_excel(excel_workbook: Workbook,
                                     json_file: Path | None = None,
                                     ) -> list[dict[str, dict[str, str]]]:
    """Создание всех inline кнопок из таблицы excel."""

    def _get_one_button_with_content(col_num: int, _ws: Worksheet) -> dict[str, dict[str, str]]:
        """Генерирует словарь с именем кнопки и её значениями."""
        if col_num + 1 >= _ws.max_column:
            return {}
        res = {}

        for _row in _ws.iter_rows(min_row=3, max_row=ws.max_row):
            values = {}
            if _row[col_num].value is not None:
                for _next_row in _ws.iter_rows(min_row=_row[col_num].row + 1, max_row=ws.max_row):
                    if (
                            _next_row[col_num + 1].value is None
                            and _next_row[col_num].value is not None
                    ):
                        break
                    if _next_row[col_num + 1].value:
                        values[_next_row[col_num + 1].coordinate] = _next_row[col_num + 1].value

                    if values:
                        res[f"{_row[col_num].coordinate}"] = {
                            "button_name": _row[col_num].value,
                            "button_values": values,
                        }
        return res

    ws = excel_workbook.active

    button_layers = [_get_one_button_with_content(col, ws) for col in range(ws.max_column - 1)]

    if json_file:
        with json_file.open("w", encoding="utf-8") as result_file:
            json.dump(button_layers, result_file, ensure_ascii=False, indent=4)

    return button_layers


def read_categories_json_file() -> list:
    """Чтение json файла с inline кнопками категорий нарушений."""
    with VIOLATION_CATEGORY_JSON_FILE.open("r", encoding="utf-8") as json_file:
        return json.load(json_file)


if __name__ == "__main__":
    _file_path = BASEDIR / "misc" / "ot_bot_inline_buttons.xlsx"
    _json_file = BASEDIR / "misc" / "inline_buttons.json"
    wb = load_workbook(filename=_file_path)
    result = create_inline_buttons_from_excel(excel_workbook=wb, json_file=_json_file)
