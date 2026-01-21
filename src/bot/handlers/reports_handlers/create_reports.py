"""Создание отчётов нарушений."""

import io
import platform
import subprocess

from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from typing import Sequence

from bot.enums import ViolationStatus
from bot.db.models import UserModel, ViolationModel
from bot.logger_config import log
from bot.handlers.reports_handlers.reports_utils import remove_default_sheet
from bot.handlers.reports_handlers.generate_typst import generate_typst
from bot.config import settings


def write_typst_file(created_by: UserModel, violations: tuple, typ_file: Path) -> None:
    typst_document = generate_typst(violations, created_by=created_by)

    with typ_file.open("w", encoding="utf-8") as tf:
        tf.write(typst_document)


def create_typst_report(created_by: UserModel, violations: Sequence[ViolationModel]) -> Path:
    """Создание отчёта pdf с помощью typst."""
    first, last= violations[0], violations[-1]
    if first == last:
        file_number = str(first.number)
        log.info("создается предписание № {f.number}({f.id})", f=first)
    else:
        file_number = f"{first.number}_{last.number}"
        log.info("создаются предписания №№ {f.number}({f.id}) - {l.number}({l.id})", f=first, l=last)



    typ_file = settings.report_typ_file
    write_typst_file(created_by, violations, typ_file)
    # pdf_file = settings.report_pdf_file
    pdf_file = settings.typst_dir / f"предписание_{file_number}.pdf"
    if pdf_file.exists():
        try:
            pdf_file.unlink()
        except OSError as e:
            log.error("Ошибка удаления файла {f} перед созданием нового", f=pdf_file)
            log.exception(e)
    if platform.system() == "Windows":
        typst_command = settings.typst_dir / "typst.exe"
        cmd = [typst_command, "compile", "--root", settings.BASE_DIR, typ_file, pdf_file]

    else:
        cmd = ["typst", "compile", "--root", settings.BASE_DIR, typ_file, pdf_file]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True,
        )
        log.info(result.stdout)
        log.warning(result.stderr)
    except subprocess.CalledProcessError as e:
        log.warning(e.stdout)
        log.error(e.stderr)
        msg = "Не удалось скомпилировать Typst файл"
        raise RuntimeError(msg)
    typ_file.unlink(missing_ok=True)
    log.success(f"PDF успешно создан: {pdf_file}")
    return pdf_file


def create_static_report(violations: tuple) -> bytes:
    """Создание статистического отчёта xlsx."""
    # TODO добавить период выгрузки для violations после получения достаточного объема данных
    wb = Workbook()
    wb = remove_default_sheet(wb)

    # полный отчёт
    wb.create_sheet(title="Полный отчёт")
    full_report_ws = wb["Полный отчёт"]
    # шапка
    full_report_ws.append(
        [
            "Номер нарушения",  # ["number"]
            "Место нарушения",  # ["area"]["name"]
            "Описание нарушения",  # ["description]
            "Ответственный",  # ["area"]["responsible_user"] if responsible_user_id else ["area"]["responsible_text"]
            "Категория нарушения",  # ["category"]
            "Мероприятия",  # ["actions_needed"]
            "Статус",  # ["status"]
            "Дата обнаружения",  # ["created_at"]
            "Дата закрытия",  # ["updated_at"] if status == <ViolationStatus.COMPLETED: 'завершено'> else ""
            "Обнаружено работником",  # ["detector"]["first_name"] + ["detector"]["user_role"]
        ]
    )
    # данные полного отчёта
    for violation in violations:
        full_report_ws.append(
            [
                violation.number,
                violation.area.name,
                violation.description,
                violation.area.responsible_user.first_name
                if violation.area.responsible_user_id
                else (violation.area.responsible_text),
                violation.category,
                violation.actions_needed,
                violation.status,
                violation.created_at.strftime("%d.%m.%Y %H:%M:%S"),
                violation.updated_at.strftime("%d.%m.%Y %H:%M:%S")
                if violation.status == ViolationStatus.CORRECTED
                else "",
                f"ФИО: {violation.detector.first_name} Роль:{violation.detector.user_role}",
            ]
        )

    # отчёт по каждому месту нарушения
    area_report_data = {}

    for violation in violations:
        area_name = violation.area.name
        responsible = (
            violation.area.responsible_user.first_name
            if violation.area.responsible_user_id
            else (violation.area.responsible_text)
        )
        status = violation.status

        if area_name not in area_report_data:
            area_report_data[area_name] = {"violations": defaultdict(int)}

        area_report_data[area_name]["violations"][status] += 1
        area_report_data[area_name]["responsible"] = responsible

    wb.create_sheet(title="Места нарушения")
    area_report_ws = wb["Места нарушения"]

    # шапка
    area_report_ws.append(
        [
            "Место нарушения",
            "Ответственный",
            ViolationStatus.ACTIVE.value,
            ViolationStatus.REVIEW.value,
            ViolationStatus.CORRECTED.value,
            ViolationStatus.REJECTED.value,
        ]
    )
    # данные полного отчёта
    for area, violation_statuses in area_report_data.items():
        area_report_ws.append(
            [
                area,
                violation_statuses["responsible"],
                violation_statuses["violations"].get(ViolationStatus.ACTIVE.value, 0),
                violation_statuses["violations"].get(ViolationStatus.REVIEW.value, 0),
                violation_statuses["violations"].get(ViolationStatus.CORRECTED.value, 0),
                violation_statuses["violations"].get(ViolationStatus.REJECTED.value, 0),
            ]
        )

    # результат
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
